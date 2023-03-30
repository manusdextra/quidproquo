"""
Quid Pro Quo

Take an Excel spreadsheet full of vocabulary and turn it into a Kahoot Quiz
"""

import argparse
import pathlib
import random

from openpyxl import load_workbook
import pandas


Word = tuple[str, str, str]
QuizItem = list[str]


def parse_args() -> argparse.Namespace:
    """Get command line arguments"""
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-s",
        "--size",
        type=int,
        help="How many questions in the quiz?",
        default=30,
    )
    return parser.parse_args()


class Config:
    """
    Collection of important variables such as source file locations
    """

    # this is necessary because the REPL doesn't have access to __file__
    try:
        root = pathlib.Path(__file__).parents[1]
    except NameError:
        root = pathlib.Path("/home/henning/code/projects/quidproquo")
    inputs = root / "inputs"
    output = root / "output"

    @property
    def files(self) -> list:
        """load all input files into memory"""
        return [pandas.ExcelFile(file) for file in self.inputs.iterdir()]

    @property
    def template(self) -> pandas.DataFrame:
        """provide kahoot template and select appropriate regions"""
        return pandas.read_excel(
            self.root / "template.xlsx",
            engine="openpyxl",
            skiprows=7,
            index_col=0,
        )


def import_sheets(config: Config) -> pandas.DataFrame:
    """
    Load files and analyse them
    """
    sourcefile = config.files[0]
    sheets = [
        pandas.read_excel(sourcefile, sheet_name=sheet_name, engine="openpyxl")
        for sheet_name in sourcefile.sheet_names
    ]
    return pandas.concat(sheets)


def filter_by_pos(
    dframe: pandas.DataFrame,
    pos: str,
) -> list[Word]:
    """
    Return a new DataFrame containing only the specified parts of speech.
    If a word has several meanings, this will split them into separate definitions,
    one for each part of speech.
    """
    all_words = dframe[(dframe["Part of speech"] == pos)][["Word", "Definition"]]
    words = []
    separator = ", "
    if separator not in pos:
        for word, definition in all_words.itertuples(index=False):
            words.append((word, pos, definition))
        return words
    pos1, pos2 = pos.split(separator)
    for word, definitions in all_words.itertuples(index=False):
        if "\n" in definitions:
            def1, def2 = definitions.split("\n")
            # cut leading numbers
            def1 = def1[3:]
            def2 = def2[3:]
        else:
            def1 = definitions
            def2 = definitions
        words.append((word, pos1, def1))
        words.append((word, pos2, def2))
    return words


def sort_by_pos(collection: list[Word]):
    sorted_words = {}
    for item in collection:
        word, pos, definition = item
        try:
            sorted_words[pos].append((word, definition))
        except KeyError:
            sorted_words[pos] = [(word, definition)]
    return sorted_words


def arrange(dframe: pandas.DataFrame):
    pos_list = list(dframe["Part of speech"].value_counts().index)
    unsorted_words = []
    for pos in pos_list:
        unsorted_words.extend(filter_by_pos(dframe, pos))
    sorted_words = sort_by_pos(unsorted_words)
    return sorted_words


def find_proportions(wordlist) -> dict[str, int]:
    """Find out which parts of speech appear most frequently
    in the source material"""
    return {pos: len(words) for pos, words in wordlist.items()}


def choose_random_words(words: dict[str, Word], size: int) -> list[QuizItem]:
    """For a randomly chosen part of speech, choose a definition, the right
    word, and then three other words from that same category"""
    questions = []
    for _ in range(0, size):
        pos = random.choice(list(words.keys()))
        right_answer, definition = random.choice(words[pos])
        wrong_answers = [
            random.choice(words[pos])[0],
            random.choice(words[pos])[0],
            random.choice(words[pos])[0],
        ]
        questions.append(
            [
                definition,
                right_answer,
                *wrong_answers,
                "20",
                "1",
            ]
        )
    return questions


def output(config, lines) -> None:
    """Check for outfile and write rows of formatted questions and
    answers to it"""
    wb = load_workbook(config.root / "template.xlsx")
    ws = wb.active
    for i, item in enumerate(lines):
        for j, item in enumerate(item):
            ws.cell(row=i + 9, column=j + 2, value=item)

    outfile = config.output / "kahoot.xlsx"
    if not outfile.exists:
        outfile.touch()
    wb.save(outfile)


if __name__ == "__main__":
    random.seed()
    args = parse_args()
    conf = Config()

    sorted_data = arrange(import_sheets(conf))

    proportions = find_proportions(sorted_data)

    quiz = choose_random_words(sorted_data, args.size)
    output(
        conf,
        quiz,
    )
